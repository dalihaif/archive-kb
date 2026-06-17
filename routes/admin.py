"""
管理后台 API
"""
import json
import hashlib
import re
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session, send_file
from models import db, MonitorSource, Policy, Category, KnowledgeItem, MonitorLog, User
from crawler import PolicyCrawler, compute_hash


# ===== 智能分类建议引擎 =====

CATEGORY_KEYWORDS = {
    # 关键词 → [(分类ID, 权重)]
    # 权重越高越优先匹配
    '档案': [(1, 10), (11, 8)],        # 档案法规 / 档案局
    '归档': [(1, 10), (11, 8)],
    '全宗': [(1, 8)],
    '文献遗产': [(1, 8), (2, 5)],
    '档案局': [(11, 15)],
    '档案馆': [(1, 10), (11, 8)],
    '档案事业': [(1, 8), (11, 8)],
    '档案管理': [(1, 10)],
    '档案保管': [(14, 12)],
    '档案法': [(5, 15)],
    '实施条例': [(6, 12)],
    '管理办法': [(6, 8)],
    '管理规定': [(6, 8)],

    '卫健委': [(10, 15)],
    '卫生': [(10, 12)],
    '医疗': [(10, 12)],
    '医院': [(10, 10)],
    '疾病': [(10, 10)],
    '诊疗': [(10, 10)],
    '医保': [(10, 10)],
    '医药': [(10, 10)],
    '健康': [(10, 8)],
    '疫情防控': [(10, 12)],
    '公共卫生': [(10, 12)],
    '护理': [(10, 8)],

    '标准': [(2, 12), (8, 8)],
    '标准化': [(2, 15)],
    'DA/T': [(8, 20)],
    'GB/T': [(2, 15)],
    'ISO': [(9, 20)],
    '行业标准': [(2, 15)],
    '国际标准': [(9, 20)],
    '规范': [(2, 8)],

    '国家法律': [(5, 20)],
    '法律': [(5, 12)],
    '中华人民共和国': [(5, 15)],
    '条例': [(6, 10)],
    '国务院': [(6, 12)],
    '行政法规': [(6, 15)],
    '地方法规': [(7, 20)],
    '省': [(7, 6)],
    '市': [(7, 6)],

    '数字化': [(13, 15)],
    '电子档案': [(13, 15)],
    '电子文件': [(13, 15)],
    '信息系统': [(13, 10)],
    '数据库': [(13, 8)],
    '人工智能': [(13, 10)],
    '信息化': [(13, 12)],
    'OCR': [(13, 10)],
    '扫描': [(13, 8)],

    '保管': [(14, 12)],
    '保护': [(14, 12)],
    '安全': [(14, 10)],
    '防汛': [(14, 8)],
    '灾备': [(14, 10)],
    '修复': [(14, 8)],

    '分类': [(12, 15)],
    '编目': [(12, 15)],
    '元数据': [(12, 12)],
    '著录': [(12, 12)],
    '主题词': [(12, 8)],

    '通知': [(6, 4), (10, 3), (11, 3)],
    '公告': [(6, 3), (10, 3), (11, 3)],
    '规定': [(6, 6)],
    '办法': [(6, 6)],
}


def suggest_category(title, source_category=''):
    """根据标题自动建议知识库分类，返回 [(category_id, name, score), ...]"""
    scores = {}  # category_id -> score
    
    # 1. 关键词匹配
    for keyword, cat_weights in CATEGORY_KEYWORDS.items():
        if keyword.lower() in title.lower():
            for cat_id, weight in cat_weights:
                scores[cat_id] = scores.get(cat_id, 0) + weight
    
    # 2. 根据来源分类加分
    source_map = {
        '卫健委': [10], '档案局': [11], '法律法规': [5, 6],
        '行业标准': [2, 8], '行业动态': [2, 3],
    }
    if source_category in source_map:
        for cat_id in source_map[source_category]:
            scores[cat_id] = scores.get(cat_id, 0) + 5
    
    # 3. 如果分数为空，给予通用分类默认低分
    if not scores:
        # 检查是否有其他可用的兜底逻辑
        if any(kw in title for kw in ['档案', '文件', '文献']):
            scores[1] = 4
        else:
            scores[3] = 2  # 默认政策文件
    
    # 4. 收集所有分类名称，按分数排序
    all_cats = {c.id: c for c in Category.query.all()}
    results = []
    for cat_id, score in sorted(scores.items(), key=lambda x: -x[1]):
        cat = all_cats.get(cat_id)
        if cat:
            parent_name = ''
            if cat.parent_id:
                parent = all_cats.get(cat.parent_id)
                if parent:
                    parent_name = parent.name
            results.append({
                'category_id': cat.id,
                'name': cat.name,
                'parent_name': parent_name,
                'full_path': f'{parent_name} > {cat.name}' if parent_name else cat.name,
                'score': score,
            })
    
    return results[:5]  # 最多返回5个建议

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def _fts_search_policies(keyword):
    """FTS5 全文搜索政策，返回匹配的 Policy ID 列表"""
    try:
        from sqlalchemy import text
        # FTS5 查询语法：用 * 做前缀匹配
        safe_kw = keyword.replace('"', '').replace("'", '')
        results = db.session.execute(
            text("SELECT rowid FROM policies_fts WHERE policies_fts MATCH :q LIMIT 500"),
            {"q": f'"{safe_kw}"'}
        ).fetchall()
        if results:
            return [r[0] for r in results]
    except Exception:
        pass
    return None


def _fts_search_knowledge(keyword):
    """FTS5 全文搜索知识库，返回匹配的 KnowledgeItem ID 列表"""
    try:
        from sqlalchemy import text
        safe_kw = keyword.replace('"', '').replace("'", '')
        results = db.session.execute(
            text("SELECT rowid FROM knowledge_fts WHERE knowledge_fts MATCH :q LIMIT 500"),
            {"q": f'"{safe_kw}"'}
        ).fetchall()
        if results:
            return [r[0] for r in results]
    except Exception:
        pass
    return None


@admin_bp.route('/login')
def login():
    """自动跳转到管理后台"""
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/logout')
def logout():
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/')
def dashboard():
    """管理后台首页"""
    policy_count = Policy.query.count()
    source_count = MonitorSource.query.count()
    knowledge_count = KnowledgeItem.query.count()
    today_policies = Policy.query.filter(Policy.created_at >= db.func.date('now')).count()
    today_logs = MonitorLog.query.filter(MonitorLog.created_at >= db.func.date('now')).count()

    last_logs = MonitorLog.query.order_by(MonitorLog.created_at.desc()).limit(10).all()

    category_stats = db.session.query(
        MonitorSource.category, db.func.count(MonitorSource.id)
    ).group_by(MonitorSource.category).all()

    return render_template('admin/dashboard.html',
        policy_count=policy_count,
        source_count=source_count,
        knowledge_count=knowledge_count,
        today_policies=today_policies,
        today_logs=today_logs,
        last_logs=last_logs,
        category_stats=category_stats)


# ===== 数据源管理 =====

@admin_bp.route('/sources')
def source_list():
    sources = MonitorSource.query.order_by(MonitorSource.category, MonitorSource.name).all()
    return render_template('admin/sources.html', sources=sources)


@admin_bp.route('/api/sources', methods=['POST'])
def source_create():
    data = request.json
    source = MonitorSource(
        name=data['name'],
        url=data['url'],
        source_type=data.get('source_type', 'web'),
        category=data.get('category', '其他'),
        crawl_method=data.get('crawl_method', 'get'),
        selectors=data.get('selectors', '{}'),
        enabled=data.get('enabled', True),
    )
    db.session.add(source)
    db.session.commit()
    return jsonify({'success': True, 'id': source.id})


@admin_bp.route('/api/sources/<int:source_id>', methods=['PUT'])
def source_update(source_id):
    source = MonitorSource.query.get_or_404(source_id)
    data = request.json
    for field in ['name', 'url', 'source_type', 'category', 'crawl_method', 'selectors', 'enabled']:
        if field in data:
            setattr(source, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/sources/<int:source_id>', methods=['DELETE'])
def source_delete(source_id):
    source = MonitorSource.query.get_or_404(source_id)
    db.session.delete(source)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/sources/<int:source_id>/crawl', methods=['POST'])
def source_crawl(source_id):
    """手动触发单个数据源抓取"""
    source = MonitorSource.query.get_or_404(source_id)
    crawler = PolicyCrawler(db.session)
    crawler.crawl_source(source)
    return jsonify({'success': True, 'status': source.last_crawl_status})


# ===== 政策管理 =====

@admin_bp.route('/policies')
def policy_list():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('q', '')
    category = request.args.get('category', '')

    query = Policy.query
    if category:
        source_ids = [s.id for s in MonitorSource.query.filter_by(category=category).all()]
        if source_ids:
            query = query.filter(Policy.source_id.in_(source_ids))
    if keyword:
        fts_ids = _fts_search_policies(keyword)
        if fts_ids:
            query = query.filter(Policy.id.in_(fts_ids))
        else:
            query = query.filter(
                db.or_(Policy.title.contains(keyword), Policy.summary.contains(keyword),
                       Policy.content.contains(keyword), Policy.tags.contains(keyword))
            )

    query = query.order_by(Policy.created_at.desc())
    pagination = query.paginate(page=page, per_page=20, error_out=False)

    categories = db.session.query(MonitorSource.category).distinct().all()
    category_list = [c[0] for c in categories if c[0]]

    all_categories = Category.query.order_by(Category.sort_order).all()
    return render_template('admin/policies.html',
        policies=pagination.items, pagination=pagination,
        categories=category_list, current_category=category, keyword=keyword,
        all_categories=all_categories)


@admin_bp.route('/api/policies/<int:policy_id>', methods=['GET'])
def policy_get(policy_id):
    """获取单条政策详情"""
    policy = Policy.query.get_or_404(policy_id)
    return jsonify({
        'success': True,
        'policy': {
            'id': policy.id,
            'title': policy.title,
            'summary': policy.summary or '',
            'content': policy.content or '',
            'tags': policy.tags or '',
            'file_type': policy.file_type or '',
            'authority': policy.authority or '',
            'pub_date': str(policy.pub_date) if policy.pub_date else '',
            'source_url': policy.source_url or '',
        }
    })


@admin_bp.route('/api/policies/batch', methods=['PUT'])
def policy_batch_update():
    """批量修改政策"""
    data = request.json
    item_ids = data.get('ids', [])
    updates = data.get('updates', {})

    if not item_ids or not updates:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    allowed = ['file_type', 'tags', 'title', 'summary', 'content', 'is_pinned']
    update_data = {k: v for k, v in updates.items() if k in allowed}

    updated_count = Policy.query.filter(Policy.id.in_(item_ids)).update(
        update_data, synchronize_session=False
    )
    db.session.commit()
    return jsonify({'success': True, 'updated': updated_count})


@admin_bp.route('/api/policies/<int:policy_id>', methods=['PUT'])
def policy_update(policy_id):
    policy = Policy.query.get_or_404(policy_id)
    data = request.json
    for field in ['title', 'summary', 'content', 'tags', 'is_pinned', 'file_type']:
        if field in data:
            setattr(policy, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/policies/<int:policy_id>', methods=['DELETE'])
def policy_delete(policy_id):
    policy = Policy.query.get_or_404(policy_id)
    db.session.delete(policy)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/policies/batch-delete', methods=['POST'])
def policy_batch_delete():
    ids = request.json.get('ids', [])
    Policy.query.filter(Policy.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/policies/<int:policy_id>/suggest-category', methods=['GET'])
def policy_suggest_category(policy_id):
    """根据标题智能建议分类"""
    policy = Policy.query.get_or_404(policy_id)
    source_cat = policy.source.category if policy.source else ''
    suggestions = suggest_category(policy.title, source_cat)
    return jsonify({'success': True, 'title': policy.title, 'suggestions': suggestions})


@admin_bp.route('/api/policies/batch-suggest', methods=['POST'])
def policy_batch_suggest():
    """批量智能建议分类，返回按分类聚合的建议"""
    ids = request.json.get('ids', [])
    titles = request.json.get('titles', [])
    
    all_suggestions = {}
    for pid in ids:
        policy = Policy.query.get(pid)
        if policy:
            source_cat = policy.source.category if policy.source else ''
            suggestions = suggest_category(policy.title, source_cat)
            all_suggestions[str(pid)] = suggestions
    
    # 汇总：所有条目最常被建议的分类
    cat_counts = {}
    for pid, suglist in all_suggestions.items():
        for s in suglist[:2]:  # 每个条目取前2个建议
            cid = s['category_id']
            if cid not in cat_counts:
                cat_counts[cid] = {'count': 0, 'name': s['name'], 'parent_name': s['parent_name'],
                                   'full_path': s['full_path'], 'category_id': cid}
            cat_counts[cid]['count'] += 1
    
    ranked = sorted(cat_counts.values(), key=lambda x: -x['count'])
    
    # 每个条目的最佳单独建议
    per_item = {}
    all_cats = {c.id: c for c in Category.query.all()}
    for pid, suglist in all_suggestions.items():
        if suglist:
            per_item[str(pid)] = suglist[0]
    
    return jsonify({
        'success': True,
        'ranked_suggestions': ranked[:8],
        'per_item': per_item,
        'total_items': len(ids),
    })


@admin_bp.route('/api/policies/batch-to-knowledge', methods=['POST'])
def policy_batch_to_knowledge():
    """批量将政策转入知识库"""
    items = request.json.get('items', [])
    # items: [{policy_id: 1, category_id: 5}, ...]
    
    count = 0
    for item_data in items:
        policy = Policy.query.get(item_data['policy_id'])
        if not policy:
            continue
        category_id = item_data.get('category_id')
        if not category_id:
            continue
        
        # 检查是否已转入（按标题+分类去重）
        existing = KnowledgeItem.query.filter_by(
            title=policy.title, category_id=category_id
        ).first()
        if existing:
            continue
        
        ki = KnowledgeItem(
            category_id=category_id,
            title=policy.title,
            content=policy.content or policy.summary or '',
            source=policy.authority or '自动采集',
            source_url=policy.url,
            tags=policy.tags,
        )
        db.session.add(ki)
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count})


@admin_bp.route('/api/policies/to-knowledge', methods=['POST'])
def policy_to_knowledge():
    """将单条政策条目转为知识库条目"""
    policy_id = request.json.get('policy_id')
    category_id = request.json.get('category_id')
    policy = Policy.query.get_or_404(policy_id)

    item = KnowledgeItem(
        category_id=category_id,
        title=policy.title,
        content=policy.content or policy.summary or '',
        source=policy.authority or '自动采集',
        source_url=policy.url,
        tags=policy.tags,
    )
    db.session.add(item)
    db.session.commit()
    return jsonify({'success': True, 'id': item.id})


# ===== 分类管理 =====

@admin_bp.route('/categories')
def category_list():
    categories = Category.query.order_by(Category.sort_order).all()
    return render_template('admin/categories.html', categories=categories)


@admin_bp.route('/api/categories', methods=['POST'])
def category_create():
    data = request.json
    cat = Category(
        name=data['name'],
        parent_id=data.get('parent_id'),
        sort_order=data.get('sort_order', 0),
        icon=data.get('icon', 'bi-folder'),
        description=data.get('description', ''),
    )
    db.session.add(cat)
    db.session.commit()
    return jsonify({'success': True, 'id': cat.id})


@admin_bp.route('/api/categories/<int:cat_id>', methods=['PUT'])
def category_update(cat_id):
    cat = Category.query.get_or_404(cat_id)
    data = request.json
    for field in ['name', 'parent_id', 'sort_order', 'icon', 'description']:
        if field in data:
            setattr(cat, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/categories/<int:cat_id>', methods=['DELETE'])
def category_delete(cat_id):
    cat = Category.query.get_or_404(cat_id)
    if cat.children:
        return jsonify({'success': False, 'message': '该分类下有子分类，不能删除'}), 400
    db.session.delete(cat)
    db.session.commit()
    return jsonify({'success': True})


# ===== 知识库管理 =====

def _get_category_ids(cat_id):
    """获取父分类及所有子分类的ID列表（用于筛选时包含子分类）"""
    cat_ids = [cat_id]
    for child in Category.query.filter_by(parent_id=cat_id).all():
        cat_ids.append(child.id)
    return cat_ids


def _get_category_total_count(cat):
    """获取分类的条目总数（包含子分类）"""
    if cat.children:
        cat_ids = [cat.id] + [c.id for c in cat.children]
        return KnowledgeItem.query.filter(KnowledgeItem.category_id.in_(cat_ids)).count()
    return cat.knowledge_items.count()


@admin_bp.route('/knowledge')
def knowledge_list():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('q', '')
    cat_id = request.args.get('category_id', 0, type=int)

    query = KnowledgeItem.query
    if cat_id:
        cat_ids = _get_category_ids(cat_id)
        query = query.filter(KnowledgeItem.category_id.in_(cat_ids))
    if keyword:
        fts_ids = _fts_search_knowledge(keyword)
        if fts_ids:
            query = query.filter(KnowledgeItem.id.in_(fts_ids))
        else:
            query = query.filter(
                db.or_(KnowledgeItem.title.contains(keyword), KnowledgeItem.content.contains(keyword))
            )
    query = query.order_by(KnowledgeItem.updated_at.desc())
    pagination = query.paginate(page=page, per_page=20, error_out=False)

    categories = Category.query.order_by(Category.sort_order).all()
    total_all = KnowledgeItem.query.count()  # 全局总条数

    # 构建带计数的分类列表（父分类累加子分类）
    cat_counts = {}
    for c in categories:
        cat_counts[c.id] = _get_category_total_count(c)

    return render_template('admin/knowledge.html',
        items=pagination.items, pagination=pagination,
        categories=categories, current_category_id=cat_id, keyword=keyword,
        total_all=total_all, cat_counts=cat_counts)


@admin_bp.route('/api/knowledge/<int:item_id>', methods=['GET'])
def knowledge_get(item_id):
    """获取单条知识条目详情，用于编辑时回填表单"""
    item = KnowledgeItem.query.get_or_404(item_id)
    return jsonify({
        'success': True,
        'item': {
            'id': item.id,
            'category_id': item.category_id,
            'title': item.title,
            'content': item.content or '',
            'source': item.source or '',
            'source_url': item.source_url or '',
            'tags': item.tags or '',
            'is_pinned': item.is_pinned,
        }
    })


@admin_bp.route('/api/knowledge/batch', methods=['PUT'])
def knowledge_batch_update():
    """批量修改知识条目"""
    data = request.json
    item_ids = data.get('ids', [])
    updates = data.get('updates', {})

    if not item_ids or not updates:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    allowed = ['category_id', 'title', 'content', 'source', 'source_url', 'tags', 'is_pinned']
    update_data = {k: v for k, v in updates.items() if k in allowed}

    updated_count = KnowledgeItem.query.filter(KnowledgeItem.id.in_(item_ids)).update(
        update_data, synchronize_session=False
    )
    db.session.commit()
    return jsonify({'success': True, 'updated': updated_count})


@admin_bp.route('/api/knowledge', methods=['POST'])
def knowledge_create():
    data = request.json
    item = KnowledgeItem(
        category_id=data['category_id'],
        title=data['title'],
        content=data.get('content', ''),
        source=data.get('source', ''),
        source_url=data.get('source_url', ''),
        tags=data.get('tags', ''),
        is_pinned=data.get('is_pinned', False),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify({'success': True, 'id': item.id})


@admin_bp.route('/api/knowledge/<int:item_id>', methods=['PUT'])
def knowledge_update(item_id):
    item = KnowledgeItem.query.get_or_404(item_id)
    data = request.json
    for field in ['category_id', 'title', 'content', 'source', 'source_url', 'tags', 'is_pinned']:
        if field in data:
            setattr(item, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/knowledge/<int:item_id>', methods=['DELETE'])
def knowledge_delete(item_id):
    item = KnowledgeItem.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/knowledge/batch-suggest', methods=['POST'])
def knowledge_batch_suggest():
    """知识库批量智能分类建议"""
    ids = request.json.get('ids', [])
    per_item = {}
    for kid in ids:
        item = KnowledgeItem.query.get(kid)
        if item:
            suggestions = suggest_category(item.title, item.source or '')
            if suggestions:
                per_item[str(kid)] = suggestions[0]
    return jsonify({'success': True, 'suggestions': per_item})


# ===== 全局抓取 =====

@admin_bp.route('/api/crawl-all', methods=['POST'])
def crawl_all():
    crawler = PolicyCrawler(db.session)
    crawler.crawl_all_enabled()
    return jsonify({'success': True})


# ===== 统计API =====

@admin_bp.route('/api/stats')
def api_stats():
    return jsonify({
        'policy_count': Policy.query.count(),
        'source_count': MonitorSource.query.count(),
        'knowledge_count': KnowledgeItem.query.count(),
        'today_policies': Policy.query.filter(Policy.created_at >= db.func.date('now')).count(),
    })


# ===== Excel 导出 =====

@admin_bp.route('/api/policies/export')
def policy_export():
    """导出政策为 Excel"""
    from datetime import date, datetime
    try:
        import openpyxl
    except ImportError:
        return jsonify({'success': False, 'error': '请先安装 openpyxl: pip install openpyxl'}), 500

    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    category = request.args.get('category', '')
    keyword = request.args.get('q', '')

    query = Policy.query
    if date_from:
        try:
            query = query.filter(Policy.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Policy.created_at <= datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass
    if category:
        source_ids = [s.id for s in MonitorSource.query.filter_by(category=category).all() if s]
        if source_ids:
            query = query.filter(Policy.source_id.in_(source_ids))
    if keyword:
        fts_ids = _fts_search_policies(keyword)
        if fts_ids:
            query = query.filter(Policy.id.in_(fts_ids))
        else:
            query = query.filter(Policy.title.contains(keyword))

    policies = query.order_by(Policy.created_at.desc()).limit(5000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '政策清单'

    # 表头
    headers = ['序号', '标题', '类型', '发布机构', '发布日期', '来源', '标签', '摘要', '链接']
    header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    header_fill = openpyxl.styles.PatternFill(start_color='1A2A4A', end_color='1A2A4A', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, p in enumerate(policies, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=p.title)
        ws.cell(row=i+1, column=3, value=p.file_type or '')
        ws.cell(row=i+1, column=4, value=p.authority or '')
        ws.cell(row=i+1, column=5, value=str(p.pub_date) if p.pub_date else '')
        ws.cell(row=i+1, column=6, value=p.source.name if p.source else '')
        ws.cell(row=i+1, column=7, value=p.tags or '')
        ws.cell(row=i+1, column=8, value=(p.summary or '')[:500])
        ws.cell(row=i+1, column=9, value=p.url or '')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 40

    import io, os
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'政策清单_{date.today().strftime("%Y%m%d")}.xlsx'
    )


# ===== 数据备份/恢复 =====

@admin_bp.route('/api/backup')
def backup_database():
    """下载数据库备份"""
    import os
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'archive.db')
    if not os.path.exists(db_path):
        return jsonify({'success': False, 'error': '数据库文件不存在'}), 404
    from datetime import datetime
    return send_file(
        db_path,
        mimetype='application/octet-stream',
        as_attachment=True,
        download_name=f'archive_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
    )


@admin_bp.route('/api/restore', methods=['POST'])
def restore_database():
    """恢复数据库（上传 .db 文件）"""
    import os, sqlite3
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传 .db 文件'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.endswith('.db'):
        return jsonify({'success': False, 'error': '仅支持 .db 格式的 SQLite 数据库文件'}), 400

    # 验证上传的文件是否为有效的 SQLite 数据库
    try:
        import tempfile, shutil
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.db')
        os.close(tmp_fd)
        file.save(tmp_path)
        conn = sqlite3.connect(tmp_path)
        conn.execute('SELECT count(*) FROM sqlite_master')
        conn.close()
    except Exception:
        os.unlink(tmp_path)
        return jsonify({'success': False, 'error': '上传的文件不是有效的 SQLite 数据库'}), 400

    # 替换当前数据库
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'archive.db')
    # 先备份当前数据库
    backup_path = db_path + '.before_restore'
    if os.path.exists(db_path):
        shutil.copy2(db_path, backup_path)

    shutil.move(tmp_path, db_path)
    # 重建 FTS 索引
    from models import init_fts
    init_fts()

    return jsonify({
        'success': True,
        'message': '数据库恢复成功！FTS 索引已重建。请刷新页面查看最新数据。'
    })


# ===== 关键词订阅与通知 =====

@admin_bp.route('/api/alerts/keywords', methods=['GET'])
def get_alert_keywords():
    """获取当前关键词订阅"""
    user = User.query.filter_by(username='admin').first()
    keywords = (user.alert_keywords or '') if user else ''
    return jsonify({'success': True, 'keywords': keywords})


@admin_bp.route('/api/alerts/keywords', methods=['PUT'])
def update_alert_keywords():
    """更新关键词订阅"""
    data = request.json
    keywords = data.get('keywords', '')
    user = User.query.filter_by(username='admin').first()
    if user:
        user.alert_keywords = keywords
        db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/alerts/check')
def check_alerts():
    """检查新政策中命中关键词的数量"""
    user = User.query.filter_by(username='admin').first()
    if not user or not user.alert_keywords:
        return jsonify({'success': True, 'count': 0, 'matched': []})

    keywords = [kw.strip() for kw in user.alert_keywords.split(',') if kw.strip()]
    if not keywords:
        return jsonify({'success': True, 'count': 0, 'matched': []})

    # 今天的新政策
    today_policies = Policy.query.filter(
        Policy.created_at >= db.func.date('now')
    ).order_by(Policy.created_at.desc()).all()

    matched = []
    for p in today_policies:
        title_lower = p.title.lower()
        for kw in keywords:
            if kw.lower() in title_lower:
                matched.append({
                    'id': p.id, 'title': p.title, 'keyword': kw,
                    'date': str(p.created_at.date()) if p.created_at else ''
                })
                break  # 每个政策只报一次

    return jsonify({'success': True, 'count': len(matched), 'matched': matched})
